import openpyxl
import pyperclip
import hashlib
import sys
def register():
    new_username = input('please insert here your new username > ')
    print(f'your username is: {new_username}')
    new_password = input('please insert here your new password > ')
    new_password_repeat = input('please repeat your password > ')
    if new_password == new_password_repeat:
        print('the system is currently registering your login credentials, please wait')
        try:
            wb = openpyxl.load_workbook('masterstore.xlsx')
            sheet1 = wb['Sheet1']
            ws = wb.active
            registered_users = int(sheet1.max_row)
            current_row = registered_users + 1
            ws.cell(row=current_row, column=1, value = new_username)
            ws.cell(row=current_row, column=2, value = new_password)
            wb.save('masterstore.xlsx')
            print('saved')
            #it should create a new cell where row is current_row
            # todo every time a new row of cells is created the value should be newvalue += 1
        except IOError:
            wb = openpyxl.Workbook()
            ws = wb.active
            wb.create_sheet('Sheet1', 0)
            sheet1 = wb['Sheet1']
            sheet1['A1'] = new_username
            sheet1['B1'] = new_password
            wb.save('masterstore.xlsx') #works as intended until here   
            
    else: 
        print('problem new passwords are different, please repeat the registration process')
        register()
        # output: inform user if the process had any problems and store user passwords in an excel document
register()