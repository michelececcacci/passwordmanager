import openpyxl
import pyperclip
import hashlib
import sys
# the user input and launches the corresponding function
def main():
    while True:
        first_input = input('''welcome to passwordmanager v0.0.0, press 1 if you are already registered, 
        2 if you want to register and 3 if you want to quit >''')
        if first_input == '1':
            already_registered()
        if first_input == '2':
            register()
        if first_input == '3':
            sys.exit()

def register():
        new_username = input('please insert here your new username > ')
        wb = openpyxl.load_workbook('masterstore.xlsx')
        sheet1 = wb['Sheet1']
        ws = wb.active
        registered_users = int(sheet1.max_row)
        current_row = registered_users + 1
        if new_username != '':
            for i in range(1, registered_users):
                if new_username != sheet1.cell(row=i, column=1).value:
                    pass
                else:
                    print(f'username {new_username} already used, please select another one')
                    register()
            print(f'your username is: {new_username}')
            new_password = input('please insert here your new password > ')
            new_password_repeat = input('please repeat your password > ')
            if new_password == new_password_repeat:
                print('the system is currently registering your login credentials')            
                try:
                    ws.cell(row=current_row, column=1, value = new_username)
                    ws.cell(row=current_row, column=2, value = new_password)
                    wb.save('masterstore.xlsx')
                    print('password saved')
                except IOError:
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    wb.create_sheet('Sheet1', 0)
                    sheet1 = wb['Sheet1']
                    sheet1['A1'] = new_username
                    sheet1['B1'] = new_password
                    wb.save('masterstore.xlsx') 
                    print('password saved')        
            else: 
                print('Error: new passwords are different, please repeat the registration process\n')
                register()
                #output: inform user if the process had any problems and store user passwords in an excel document
        else:
            print('a blank username is not valid, sorry')
            register()  
def already_registered():
    input_username = input('please type here your username >')
    wb = openpyxl.load_workbook('masterstore.xlsx')
    sheet1 = wb['Sheet1']
    for i in range(1, int(sheet1.max_row)+1):
        if (sheet1.cell(row=i, column=1).value) == input_username:
            input_password = input('please type here your password >')
            if sheet1.cell(row=i, column=2).value == input_password:
                print('you have logged in correctly')
                password_manager(input_username)
    print('looks like you are not already registered yet, you are being redirected to the registration process')
    register() #this section works as intended

def password_manager(user): 
    try:
        wb = openpyxl.load_workbook(f'{user}passwords.xlsx')
        sheet1 = wb['Sheet1']
        user_choiche = input('please press 1 to get a saved password, or 2 to register a new one, or 3 to delete a password:>')
        if user_choiche == '1':
            print("these are all the websites you are registered on")
            for i in range(1, int(sheet1.max_row)):
                print(sheet1.cell(row=i, column=1).value)
            sitename = input('insert the website that you want your password copied from >') 
            for i in range(1, int(sheet1.max_row)+1):
                if sheet1.cell(row=i, column=1).value == sitename:
                    pyperclip.copy(sheet1.cell(row=i, column=2).value)
                    print('password copied in clipboard')   
                    sys.exit()                             
            print("password not saved currently")
        elif user_choiche == '2':
            website_input = input('please insert  the website you would like to register your password from >')
            for i in range(2, int(sheet1.max_row+1)):
                if sheet1.cell(row=i, column=1).value == website_input:
                    print(f'password for website already registered, copying it to clipboard: {sheet1.cell(row=i, column=1).value}')
                    pyperclip.copy(sheet1.cell(row=i, column=2).value)
                    sys.exit()
                else:
                    print('password not registered') #todo this section of code doesn't let the user register a password
                    sheet1.cell(row=i, column=1).value = website_input
                    website_password =  input(f"insert the password for {website_input}") 
                    sheet1.cell(row=i, column=2).value = website_password 
                    wb.save(f'{user}passwords.xlsx') 
                    sys.exit()
        elif user_choiche == '3':
            clear_website = input('please insert the website to clear password for >')
            password_manager(user)
            for i in range(2, int(sheet1.max_row)+1):
                if (sheet1.cell(row=i, column=1).value) == clear_website:
                    sheet1.cell(row=i, column=1, value = '')
                    sheet1.cell(row=i, column=2, value = '')
                    print('passwords cleared correctly')
                sys.exit()
        website_input = input('please input again the website you want to register on >')
        password_input = input(f'please set a password for {website_input} > ')
        sheet1.cell(row=int(sheet1.max_row)+1, column=1, value=website_input)
        sheet1.cell(row=int(sheet1.max_row), column=2, value= password_input)
        wb.save(f'{user}passwords.xlsx') #todo user should be able to change his passwords
        sys.exit()
    except IOError:
        wb = openpyxl.Workbook()
        ws = wb.active
        sheet1 = wb.create_sheet('Sheet1', 0)
        wb.save(f'{user}passwords.xlsx') 
        password_manager(user)#this should work as intended
    
    #function checks if the user has already a file, otherwise it's created
    #then the user can manage passwords and input a new one from there
main()
#todo maybe try to use hash to store the passwords in a more secure way, or encrypt the files or something else